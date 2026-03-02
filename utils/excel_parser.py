import re
from datetime import datetime
from typing import Dict, List, Tuple

import openpyxl


def _fmt(val) -> str:
    """Normalize a cell value to a stripped string."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M")
    return str(val).strip()


def get_sheet_names(filepath: str) -> List[str]:
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def load_sheet_rows(
    filepath: str, sheet_name: str
) -> Tuple[List[str], List[Dict[str, str]]]:
    """Return (headers, rows) for the given sheet. Rows are dicts keyed by header."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' in '{filepath}' has no header row.")

    headers = [_fmt(h) for h in header_row]

    rows: List[Dict[str, str]] = []
    for raw_row in rows_iter:
        row_dict: Dict[str, str] = {}
        for i, val in enumerate(raw_row):
            if i < len(headers) and headers[i]:
                row_dict[headers[i]] = _fmt(val)
        # Skip completely blank rows
        if any(v for v in row_dict.values()):
            rows.append(row_dict)

    wb.close()
    return headers, rows


def find_supplier_column(headers: List[str]) -> str:
    for h in headers:
        if re.search(r"\bsupplier\b", h, re.IGNORECASE):
            return h
    raise ValueError(f"No 'Supplier' column found. Headers: {headers}")


def find_pn_column(headers: List[str]) -> str:
    for h in headers:
        if re.search(r"part\s+number\s+vp", h, re.IGNORECASE):
            return h
    raise ValueError(f"No 'Part Number VP' column found. Headers: {headers}")


def search_rows(
    rows: List[Dict[str, str]],
    supplier_col: str,
    pn_col: str,
    vendors: List[str],
    part_number: str,
) -> List[Dict[str, str]]:
    """
    Filter rows where supplier matches any of vendors (substring, case-insensitive)
    AND part_number matches Part Number VP (substring, case-insensitive).
    Empty vendors list or empty part_number skips that filter.

    Cable-length variants (the last dash-separated segment) are treated as
    equivalent: searching FT-72SMR-M8FM8FBW-ZP34-M64 also finds rows containing
    FT-72SMR-M8FM8FBW-ZP34-MXX because the length designator is not a qualifying
    factor for FAI purposes.
    """
    vendor_lowers = [v.lower() for v in vendors if v]
    pn_lower = part_number.lower()

    # Base prefix strips the last segment (cable length) so length variants all match.
    # e.g. "ft-72smr-m8fm8fbw-zp34-m64" → pn_base_prefix = "ft-72smr-m8fm8fbw-zp34-"
    # Only applied for 5+ segment PNs: for shorter PNs the last segment is part of
    # the product spec (e.g. ZP34 breakout), not the length, so stripping it would
    # match across unrelated configurations (e.g. ZP64 records for a ZP34 search).
    pn_segments = [s for s in pn_lower.split("-") if s] if part_number else []
    if len(pn_segments) >= 5:
        pn_base_prefix = "-".join(pn_segments[:-1]) + "-"
    else:
        pn_base_prefix = None

    results = []
    for row in rows:
        supplier_val = row.get(supplier_col, "").lower()
        pn_val = row.get(pn_col, "").lower()

        vendor_match = not vendor_lowers or any(vl in supplier_val for vl in vendor_lowers)
        pn_match = (
            not part_number
            or (pn_lower in pn_val)
            or (pn_base_prefix is not None and pn_base_prefix in pn_val)
        )

        if vendor_match and pn_match:
            results.append(row)

    return results


def get_suggestions_for_sheet(filepath: str, sheet_name: str) -> dict:
    """Return sorted unique vendor names, part numbers, and vendor→PNs map for the given sheet."""
    headers, rows = load_sheet_rows(filepath, sheet_name)

    supplier_col = find_supplier_column(headers)
    pn_col = find_pn_column(headers)

    vendors = sorted(set(row[supplier_col] for row in rows if row.get(supplier_col)))
    pns = sorted(set(row[pn_col] for row in rows if row.get(pn_col)))

    vendor_pns: Dict[str, List[str]] = {}
    for row in rows:
        v = row.get(supplier_col, "")
        p = row.get(pn_col, "")
        if v and p:
            if v not in vendor_pns:
                vendor_pns[v] = []
            if p not in vendor_pns[v]:
                vendor_pns[v].append(p)
    for v in vendor_pns:
        vendor_pns[v].sort()

    return {"vendors": vendors, "pns": pns, "vendor_pns": vendor_pns}
